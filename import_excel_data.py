#!/usr/bin/env python3
"""
Import data from Excel file into SQLite database
"""

import sqlite3
import openpyxl
import os
from datetime import datetime

def import_excel_data():
    """Import data from SakeRecipeDataBase.xlsx into SQLite database"""
    
    # Check if Excel file exists
    excel_file = '../../SakeRecipeDataBase.xlsx'
    if not os.path.exists(excel_file):
        print(f"Excel file not found: {excel_file}")
        return
    
    # Connect to database
    conn = sqlite3.connect('sake_recipe_db.sqlite')
    cursor = conn.cursor()
    
    # Load Excel workbook
    wb = openpyxl.load_workbook(excel_file)
    
    # Import Ingredients data
    if 'Ingredients' in wb.sheetnames:
        print("Importing Ingredients...")
        sheet = wb['Ingredients']
        
        # Skip header row
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[0] != 'ingredientID' and row[1]:  # Skip empty rows, header, and rows without type
                cursor.execute('''
                    INSERT OR REPLACE INTO ingredients 
                    (ingredientID, ingredient_type, acc_date, source, description)
                    VALUES (?, ?, ?, ?, ?)
                ''', (row[0], row[1], row[2], row[3], row[4]))
    
    # Import Recipe data
    if 'Recipe' in wb.sheetnames:
        print("Importing Recipe...")
        sheet = wb['Recipe']
        
        # Get column headers
        headers = [cell.value for cell in sheet[1]]
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[2] and row[2] != 'batchID':  # Skip empty rows and header
                # Map Excel columns to database columns
                data = {
                    'start_date': row[0],
                    'pouch_date': row[1], 
                    'batchID': row[2],
                    'batch': row[3],
                    'style': row[4],
                    'kake': row[5],
                    'koji': row[6],
                    'yeast': row[7],
                    'starter': row[8],
                    'water_type': row[9]
                }
                
                cursor.execute('''
                    INSERT OR REPLACE INTO recipe 
                    (start_date, pouch_date, batchID, batch, style, kake, koji, yeast, starter, water_type)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (data['start_date'], data['pouch_date'], data['batchID'], 
                      data['batch'], data['style'], data['kake'], data['koji'], 
                      data['yeast'], data['starter'], data['water_type']))
    
    # Import Starters data
    if 'Starters' in wb.sheetnames:
        print("Importing Starters...")
        sheet = wb['Starters']
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[2] and row[2] != 'BatchID':  # Skip empty rows and header
                cursor.execute('''
                    INSERT OR REPLACE INTO starters 
                    (date, starter_batch, batchID, amt_kake, amt_koji, amt_water, 
                     water_type, kake, koji, yeast)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (row[0], row[1], row[2], row[3], row[4], row[5], 
                      row[6], row[7], row[8], row[9]))
    
    # Import PublishNotes data
    if 'PublishNotes' in wb.sheetnames:
        print("Importing PublishNotes...")
        sheet = wb['PublishNotes']
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[0] != 'BatchID':  # Skip empty rows and header
                cursor.execute('''
                    INSERT OR REPLACE INTO publish_notes 
                    (batchID, pouch_date, style, water, abv, smv, batch_size_l, rice, description)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (row[0], row[1], row[2], row[3], row[4], row[5], 
                      row[6], row[7], row[8]))
    
    # Import Formulas data (sample data for now)
    if 'Formulas' in wb.sheetnames:
        print("Importing Formulas...")
        sheet = wb['Formulas']
        
        for row in sheet.iter_rows(min_row=3, values_only=True):  # Start from row 3
            if row[0] and isinstance(row[0], (int, float)):  # Check if first column has numeric data
                # Convert any complex objects to strings or None
                clean_row = []
                for i, val in enumerate(row[:7]):  # Only take first 7 columns
                    if val is None:
                        clean_row.append(None)
                    elif hasattr(val, '__class__') and 'ArrayFormula' in str(type(val)):
                        clean_row.append(None)  # Skip formula objects
                    elif isinstance(val, (int, float)):
                        clean_row.append(val)
                    else:
                        clean_row.append(None)
                
                cursor.execute('''
                    INSERT INTO formulas 
                    (measure_temp_c, measured_sg, measured_brix, final_sg, abv, smv, brix)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', tuple(clean_row))
    
    # Commit and close
    conn.commit()
    conn.close()
    
    print("Data import completed successfully!")

if __name__ == "__main__":
    import_excel_data()
